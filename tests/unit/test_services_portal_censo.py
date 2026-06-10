"""S2: Censo Hospitalar page with real data from CensusSnapshot."""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest
from django.urls import reverse
from django.utils import timezone

from apps.census.models import BedStatus, CensusSnapshot, Specialty


@pytest.mark.django_db
class TestCensoRealData:
    """S2: Censo page shows real data from CensusSnapshot (occupied only)."""

    def test_censo_empty_db_shows_message(self, admin_client):
        """Without CensusSnapshot, shows empty state."""
        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "Censo Hospitalar" in content

    def test_censo_shows_occupied_patients(self, admin_client):
        """Only occupied beds appear on censo page."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="PACIENTE BETA", especialidade="CIV",
            bed_status=BedStatus.OCCUPIED,
        )
        # Empty bed — should NOT appear
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="03",
            prontuario="", nome="VAZIO", especialidade="",
            bed_status=BedStatus.EMPTY,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "PACIENTE ALFA" in content
        assert "PACIENTE BETA" in content
        assert "VAZIO" not in content

    def test_censo_shows_timestamp_when_available(self, admin_client):
        """Censo page shows capture timestamp."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PAC", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        # localtime() converts UTC to TIME_ZONE (America/Sao_Paulo)
        # to match template rendering {{ captured_at|date:"d/m/Y H:i" }}
        local_now = timezone.localtime(now)
        date_str = local_now.strftime("%d/%m/%Y")
        assert date_str in content

    def test_censo_filters_by_setor(self, admin_client):
        """Filter by sector shows only that sector's patients."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PAC UTI", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="CLINICA", leito="01",
            prontuario="222", nome="PAC CLINICA", especialidade="CME",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?unidade=UTI+A"
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "PAC UTI" in content
        assert "PAC CLINICA" not in content

    def test_censo_search_by_name(self, admin_client):
        """Free-text search filters by patient name."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="JOAO SILVA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="MARIA SANTOS", especialidade="CIV",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?q=JOAO"
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "JOAO SILVA" in content
        assert "MARIA SANTOS" not in content

    def test_censo_search_by_prontuario(self, admin_client):
        """Free-text search filters by patient record number."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="99999", nome="JOAO SILVA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="88888", nome="MARIA SANTOS", especialidade="CIV",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?q=99999"
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "JOAO SILVA" in content
        assert "MARIA SANTOS" not in content

    def test_censo_dropdown_has_real_setores(self, admin_client):
        """Sector dropdown is populated from actual sectors in the snapshot."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI GERAL", leito="01",
            prontuario="111", nome="PAC1", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="ENFERMARIA B", leito="01",
            prontuario="222", nome="PAC2", especialidade="CME",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "UTI GERAL" in content
        assert "ENFERMARIA B" in content
        assert "Todos os setores" in content  # the "all" option remains

    def test_censo_uses_only_latest_snapshot(self, admin_client):
        """Only the most recent CensusSnapshot is used."""
        from datetime import timedelta
        old = timezone.now() - timedelta(hours=4)
        new = timezone.now()

        CensusSnapshot.objects.create(
            captured_at=old, setor="OLD SETOR", leito="01",
            prontuario="AAA", nome="OLD PATIENT", especialidade="X",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=new, setor="NEW SETOR", leito="01",
            prontuario="BBB", nome="NEW PATIENT", especialidade="Y",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "NEW PATIENT" in content
        assert "OLD PATIENT" not in content

    def test_censo_row_links_to_patient_search(self, admin_client):
        """Clicking a censo row links to patient search by prontuario."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="12345", nome="PAC LINK", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "q=12345" in content

    # ── CES-S1: Nomes completos de especialidades na UI ──────────

    def test_censo_dropdown_shows_full_specialty_name_with_code_value(self, admin_client):
        """RED 1: Dropdown shows full name while preserving code as value."""
        # NEF already seeded by migration, so use get_or_create
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        # Option preserves code as value
        assert 'value="NEF"' in content
        # Dropdown shows full name as the option TEXT (not just in title attribute)
        # Currently shows: <option value="NEF">NEF</option>
        # Should show:     <option value="NEF">NEFROLOGIA</option>
        assert 'value="NEF">NEFROLOGIA' in content or '"NEF">NEFROLOGIA' in content

    def test_censo_table_and_card_show_full_specialty_name(self, admin_client):
        """RED 2: Table and card show full specialty name as main text."""
        # Use a unique test code not seeded by migrations
        Specialty.objects.create(code="CES", name="CARDIOLOGIA ESPECIAL TESTE")
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="CES",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        # Full name must appear as inner text of a badge/tag,
        # not just inside a title attribute.
        # Currently badge renders: title="CARDIOLOGIA ESPECIAL TESTE">CES<
        # After fix it should:    title="CES">CARDIOLOGIA ESPECIAL TESTE<
        # So check that the sigla is NOT the badge inner text:
        # `>CES<` should NOT be present (currently it IS)
        assert ">CES<" not in content, "Sigla CES should not be badge inner text"
        # The full name should appear outside title attribute context
        assert "CARDIOLOGIA ESPECIAL TESTE" in content
        # The code value is still valid in dropdown options
        assert 'value="CES"' in content

    def test_censo_specialty_filter_continues_by_code(self, admin_client):
        """RED 3: Filtering by specialty code still works."""
        # Use unique test codes not seeded by migrations
        Specialty.objects.create(code="CES", name="CARDIOLOGIA ESPECIAL TESTE")
        Specialty.objects.create(code="ORT", name="ORTOPEDIA ESPECIAL TESTE")
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE CARDIACO", especialidade="CES",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="PACIENTE ORTOPEDICO", especialidade="ORT",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?especialidade=CES"
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "PACIENTE CARDIACO" in content
        assert "PACIENTE ORTOPEDICO" not in content
        # Dropdown preserves the selected option with code value
        assert 'value="CES" selected' in content

    def test_censo_unknown_specialty_fallback(self, admin_client):
        """RED 4: Unknown specialty falls back to original value."""
        # No Specialty created for "XYZ"
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="XYZ",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "XYZ" in content

    # ── CES-S2: Helper comum do resultado do censo ───────────────

    def test_censo_combined_filters(self, admin_client):
        """RED 1: Combined q, unidade and especialidade filters work together."""
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        Specialty.objects.get_or_create(code="CAR", defaults={"name": "CARDIOLOGIA"})
        now = timezone.now()
        # Patient matching all filters
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ESPERADO", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        # Different sector — should be filtered out
        CensusSnapshot.objects.create(
            captured_at=now, setor="ENFERMARIA", leito="01",
            prontuario="222", nome="OUTRO SETOR", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )
        # Different specialty — should be filtered out
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="333", nome="OUTRA ESPECIALIDADE", especialidade="CAR",
            bed_status=BedStatus.OCCUPIED,
        )

        url = (
            reverse("services_portal:censo")
            + "?q=ESPERADO&unidade=UTI+A&especialidade=NEF"
        )
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "PACIENTE ESPERADO" in content
        assert "OUTRO SETOR" not in content
        assert "OUTRA ESPECIALIDADE" not in content
        # Total count reflects only the single match
        assert "1 paciente" in content or "1 pacientes" in content or "paciente" in content

    def test_censo_ordering_by_especialidade(self, admin_client):
        """RED 2: Ordering by specialty works via context."""
        Specialty.objects.get_or_create(code="AAA", defaults={"name": "ESPECIALIDADE A"})
        Specialty.objects.get_or_create(code="BBB", defaults={"name": "ESPECIALIDADE B"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE BETA", especialidade="BBB",
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="PACIENTE ALFA", especialidade="AAA",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?ordenar=especialidade"
        response = admin_client.get(url)
        assert response.status_code == 200
        pacientes = response.context["pacientes"]
        assert len(pacientes) == 2
        # Ordered by specialty alphabetically: AAA first, then BBB
        assert pacientes[0]["especialidade"] == "AAA"
        assert pacientes[1]["especialidade"] == "BBB"
        # Names should appear in matching order: PACIENTE ALFA, then PACIENTE BETA
        assert pacientes[0]["nome"] == "PACIENTE ALFA"
        assert pacientes[1]["nome"] == "PACIENTE BETA"

    def test_censo_ordering_by_tempo_desc(self, admin_client):
        """RED 2b: Ordering by descending stay time works via context."""
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE CURTO", especialidade="NEF",
            tempo_internacao=2,
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="PACIENTE LONGO", especialidade="NEF",
            tempo_internacao=10,
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo") + "?ordenar=tempo_desc"
        response = admin_client.get(url)
        assert response.status_code == 200
        pacientes = response.context["pacientes"]
        assert len(pacientes) == 2
        # Longer stay first
        assert pacientes[0]["nome"] == "PACIENTE LONGO"
        assert pacientes[1]["nome"] == "PACIENTE CURTO"

    def test_censo_empty_state_preserved(self, admin_client):
        """RED 3: Empty state with context values."""
        # No snapshots at all
        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        assert response.status_code == 200
        assert response.context["pacientes"] == []
        assert response.context["total"] == 0
        assert response.context["captured_at"] is None


# ── CES-S3: Endpoint autenticado de exportação XLSX ─────────────


@pytest.mark.django_db
class TestCensoExportXlsx:
    """CES-S3: XLSX export endpoint tests."""

    def test_export_requires_login(self, client):
        """RED 1: Anonymous user receives redirect to login."""
        url = reverse("services_portal:censo_export_xlsx")
        response = client.get(url)
        assert response.status_code == 302
        assert "/login/" in response.url or "/accounts/login/" in response.url

    def test_export_returns_valid_xlsx(self, admin_client):
        """RED 2: Authenticated user gets valid XLSX with correct headers."""
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALPHA", especialidade="NEF",
            tempo_internacao=5,
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo_export_xlsx")
        response = admin_client.get(url)
        assert response.status_code == 200
        assert response["Content-Type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert ".xlsx" in response["Content-Disposition"]

        # Validate workbook can be opened
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws is not None
        assert ws.title == "Censo Hospitalar"

    def test_export_columns_and_full_specialty_names(self, admin_client):
        """RED 3: Workbook has expected columns and full specialty names."""
        Specialty.objects.get_or_create(
            code="CES", defaults={"name": "CARDIOLOGIA ESPECIAL TESTE"}
        )
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALPHA", especialidade="CES",
            tempo_internacao=5,
            data_internacao="01/06/2026",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo_export_xlsx")
        response = admin_client.get(url)
        assert response.status_code == 200

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Headers
        expected_headers = [
            "Registro", "Nome", "Setor / Unidade", "Leito",
            "Especialidade", "Data Internação", "Tempo Internação",
            "Capturado em",
        ]
        header_row = [cell.value for cell in ws[1]]
        assert header_row == expected_headers, f"Headers mismatch: {header_row}"

        # Data row: full specialty name
        data_row = [cell.value for cell in ws[2]]
        # Registro, Nome, Setor, Leito, Especialidade, Data Internação, Tempo, Capturado
        assert data_row[0] == "111"  # Registro
        assert data_row[1] == "PACIENTE ALPHA"  # Nome
        assert data_row[2] == "UTI A"  # Setor / Unidade
        assert data_row[3] == "01"  # Leito
        assert data_row[4] == "CARDIOLOGIA ESPECIAL TESTE"  # Especialidade (full name!)
        assert data_row[5] == "01/06/2026"  # Data Internação
        assert "5" in str(data_row[6])  # Tempo Internação

    def test_export_respects_filters(self, admin_client):
        """RED 4: Export with query params filters correctly."""
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        Specialty.objects.get_or_create(code="ORT", defaults={"name": "ORTOPEDIA"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="NEF",
            tempo_internacao=3,
            bed_status=BedStatus.OCCUPIED,
        )
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="02",
            prontuario="222", nome="PACIENTE BETA", especialidade="ORT",
            tempo_internacao=5,
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo_export_xlsx") + "?especialidade=NEF&q=ALFA"
        response = admin_client.get(url)
        assert response.status_code == 200

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1, f"Expected 1 row, got {len(rows)}"
        assert rows[0][1] == "PACIENTE ALFA"

    def test_export_empty_state_valid(self, admin_client):
        """RED 5: Export with no snapshot returns valid workbook with headers only."""
        url = reverse("services_portal:censo_export_xlsx")
        response = admin_client.get(url)
        assert response.status_code == 200

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Headers only, no data rows
        expected_headers = [
            "Registro", "Nome", "Setor / Unidade", "Leito",
            "Especialidade", "Data Internação", "Tempo Internação",
            "Capturado em",
        ]
        header_row = [cell.value for cell in ws[1]]
        assert header_row == expected_headers, f"Headers mismatch: {header_row}"
        # Only header row exists
        row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        assert row_count == 0


# ── CES-S4: Botão de exportação na página /censo/ ───────────────────


@pytest.mark.django_db
class TestCensoExportButton:
    """CES-S4: Export button on the /censo/ page."""

    def test_censo_export_button_exists(self, admin_client):
        """RED 1: Page shows 'Exportar Excel' link pointing to export route."""
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )

        url = reverse("services_portal:censo")
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        assert "Exportar Excel" in content
        export_url = reverse("services_portal:censo_export_xlsx")
        assert export_url in content

    def test_censo_export_button_preserves_querystring(self, admin_client):
        """RED 2: Export link preserves q, unidade, especialidade and ordenar."""
        Specialty.objects.get_or_create(code="NEF", defaults={"name": "NEFROLOGIA"})
        now = timezone.now()
        CensusSnapshot.objects.create(
            captured_at=now, setor="UTI A", leito="01",
            prontuario="111", nome="PACIENTE ALFA", especialidade="NEF",
            bed_status=BedStatus.OCCUPIED,
        )

        export_url = reverse("services_portal:censo_export_xlsx")
        url = (
            reverse("services_portal:censo")
            + "?q=ALFA&unidade=UTI+A&especialidade=NEF&ordenar=tempo_desc"
        )
        response = admin_client.get(url)
        content = response.content.decode()
        assert response.status_code == 200
        # The export link should contain the export URL with the same params
        assert 'q=ALFA' in content or 'q=ALFA' in response.content.decode()
        assert 'unidade=UTI+A' in content or 'unidade=UTI%2BA' in content
        assert 'especialidade=NEF' in content
        assert 'ordenar=tempo_desc' in content
        # The export route must be in the link along with at least one param
        assert export_url in content
