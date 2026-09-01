"""PFIF-S3: bulk patient-flow findings classifier and /censo badges.

Consolidated integration tests for the presentation classifier
(``apps.ingestion.patient_flow_findings``) and its integration on the
hospital census page (desktop row + mobile card).

Synthetic fixtures only: names are ``PACIENTE SINTEtico`` placeholders,
prontuarios are fake ranges, no legacy/browser/network access and no real
patient data. A timezone-aware synthetic ``now`` is injected everywhere.
"""

from __future__ import annotations

from dataclasses import FrozenInstanceError
from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

import openpyxl
import pytest
from django.db import connection
from django.test import Client
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from apps.census.models import BedStatus, CensusSnapshot, PatientMovement
from apps.clinical_docs.models import ClinicalEvent
from apps.ingestion.models import (
    CensusExecutionBatch,
    IngestionRun,
    IngestionRunStageMetric,
)
from apps.ingestion.patient_flow_findings import (
    ALL_FINDING_CODES,
    PatientFindingInput,
    PatientFlowFinding,
    build_patient_flow_findings,
    classify_patient_finding,
)
from apps.patients.models import Admission, Patient

BAHIA = ZoneInfo("America/Bahia")
NOW = datetime(2026, 6, 15, 12, 0, 0, tzinfo=BAHIA)

# Closed finding codes under test (R1).
R_RECENT = "recent_encounter_without_admission"
R_NEWBORN = "newborn_waiting_registration"
R_COMPANION = "possible_newborn_companion"
R_FIRST_EVOLUTION = "recent_admission_awaiting_first_evolution"
R_RESIDUAL = "suspected_legacy_residual"
R_MIRROR = "mirror_stale_admission"
ALL_CODES = {R_RECENT, R_NEWBORN, R_COMPANION, R_FIRST_EVOLUTION, R_RESIDUAL, R_MIRROR}

# Obstetric 3A source-sector identity used by rule 3.
SECTOR_3A = "3 6 - 3A - OBSTETRÍCIA CLÍNICA - HGRS"
SECTOR_3A_CODE = "654"
SECTOR_OTHER = "ENFERMARIA SINTETICA"
SECTOR_OTHER_CODE = "999"

# Expected closed presentation contract (labels are constants).
EXPECTED = {
    R_RECENT: ("Atendimento recente sem internação", "info", False),
    R_NEWBORN: ("RN aguardando registro", "info", False),
    R_COMPANION: ("Possível RN acompanhante", "warning", True),
    R_FIRST_EVOLUTION: (
        "Internação recente aguardando 1ª evolução",
        "info",
        False,
    ),
    R_RESIDUAL: (
        "Suspeita de paciente residual no legado",
        "warning",
        True,
    ),
}

MANUAL_REVIEW_MARK = "requer revisão manual"


# ── Synthetic fixture builders ───────────────────────────────────────


def make_patient(prontuario: str, *, date_of_birth=None) -> Patient:
    return Patient.objects.create(
        patient_source_key=prontuario,
        name=f"PACIENTE SINTEtico {prontuario}",
        date_of_birth=date_of_birth,
    )


def make_census_row(
    prontuario: str,
    *,
    setor: str = SECTOR_OTHER,
    setor_codigo: str = SECTOR_OTHER_CODE,
    captured_at=None,
) -> CensusSnapshot:
    return CensusSnapshot.objects.create(
        captured_at=captured_at or timezone.now(),
        setor=setor,
        setor_codigo=setor_codigo,
        leito="01A",
        prontuario=prontuario,
        nome=f"PACIENTE SINTEtico {prontuario}",
        especialidade="",
        bed_status=BedStatus.OCCUPIED,
    )


def make_recent_outcome(
    prontuario: str, *, started_at: datetime
) -> IngestionRun:
    """Synthesize a recognized recent-encounter stage outcome (PFIF-S1/S2)."""
    run = IngestionRun.objects.create(
        status="succeeded",
        intent="admissions_only",
        parameters_json={"patient_record": prontuario},
    )
    IngestionRunStageMetric.objects.create(
        run=run,
        stage_name="encounter_fallback",
        status="succeeded",
        started_at=started_at,
        finished_at=started_at,
        details_json={
            "outcome": R_RECENT,
            "recency": "recent_confirmed",
        },
    )
    return run


def make_active_admission(
    patient: Patient,
    *,
    admission_date: datetime,
    created_at: datetime,
    discharge_date: datetime | None = None,
) -> Admission:
    adm = Admission.objects.create(
        patient=patient,
        source_admission_key=f"ADM-{patient.patient_source_key}",
        admission_date=admission_date,
        discharge_date=discharge_date,
    )
    # created_at is auto_now_add; pin it deterministically for rule 1.
    Admission.objects.filter(pk=adm.pk).update(created_at=created_at)
    adm.refresh_from_db()
    return adm


def make_event(
    admission: Admission,
    *,
    happened_at: datetime,
    key: str = "evt-sintetico-1",
) -> ClinicalEvent:
    return ClinicalEvent.objects.create(
        admission=admission,
        patient=admission.patient,
        event_identity_key=key,
        content_hash=f"hash-{key}",
        happened_at=happened_at,
        author_name="DR SINTEtico",
        profession_type="medica",
        content_text="Texto clínico sintético de teste.",
    )


def seed_cohort(prefix: str, size: int) -> None:
    """Seed a mixed synthetic cohort: outcomes, newborns and admissions.

    Timestamps are wall-clock relative so the HTTP view (which injects
    ``timezone.now()``) classifies the same cohort the same way.
    """
    now = timezone.now()
    for i in range(size):
        pront = f"{prefix}{i:02d}"
        if i % 3 == 0:
            make_census_row(pront, captured_at=now)
            make_recent_outcome(pront, started_at=now - timedelta(hours=2))
        elif i % 3 == 1:
            make_patient(
                pront,
                date_of_birth=timezone.localdate() - timedelta(days=2),
            )
            make_census_row(pront, captured_at=now)
        else:
            patient = make_patient(pront)
            make_census_row(pront, captured_at=now)
            make_active_admission(
                patient,
                admission_date=now - timedelta(hours=10),
                created_at=now - timedelta(hours=10),
            )


def _query_count(ctx: CaptureQueriesContext) -> int:
    return (ctx.final_queries or 0) - (ctx.initial_queries or 0)


def patient_id_of(pront: str) -> int | None:
    # Census rows may exist without a Patient mirror (recent outcome cases).
    return (
        Patient.objects.filter(patient_source_key=pront)
        .values_list("pk", flat=True)
        .first()
    )


def make_input(prontuario, *, patient_id=None, sector="", sector_code=""):
    return PatientFindingInput(
        prontuario=prontuario,
        patient_id=patient_id,
        sector=sector,
        sector_code=sector_code,
    )


def classify_one(**kwargs):
    """Call the pure single-patient rule function with synthetic now."""
    return classify_patient_finding(now=NOW, **kwargs)


def build_bulk(inputs):
    return build_patient_flow_findings(inputs, now=NOW)


# ── RED 1: closed codes / labels / severity / review ─────────────────


class TestClosedCodesContract:
    def test_five_closed_codes_only(self):
        assert set(ALL_FINDING_CODES) == ALL_CODES

    def test_each_code_contract_with_synthetic_now(self):
        scenarios = {
            R_RECENT: dict(latest_outcome_at=NOW - timedelta(hours=2)),
            R_NEWBORN: dict(date_of_birth=(NOW - timedelta(days=2)).date()),
            R_COMPANION: dict(
                date_of_birth=(NOW - timedelta(days=6)).date(),
                sector=SECTOR_3A,
                sector_code=SECTOR_3A_CODE,
            ),
            R_FIRST_EVOLUTION: dict(
                active_admission_date=NOW - timedelta(hours=10),
                has_active_admission=True,
            ),
            R_RESIDUAL: dict(
                active_admission_date=NOW - timedelta(hours=72),
                has_active_admission=True,
            ),
        }
        for code, kwargs in scenarios.items():
            finding = classify_one(**kwargs)
            assert finding is not None, code
            assert finding.code == code
            label, severity, review = EXPECTED[code]
            assert finding.label == label
            assert finding.severity == severity
            assert finding.requires_manual_review is review

    def test_finding_dto_is_frozen_and_closed(self):
        finding = classify_one(latest_outcome_at=NOW - timedelta(hours=2))
        expected = PatientFlowFinding(
            code=R_RECENT,
            label=EXPECTED[R_RECENT][0],
            severity="info",
            requires_manual_review=False,
        )
        assert finding == expected
        with pytest.raises(FrozenInstanceError):
            finding.code = "mutated"  # type: ignore[misc]


# ── RED 2: deterministic priority (D5 order) ─────────────────────────


class TestPriorityOrder:
    def test_recent_outcome_beats_newborn(self):
        finding = classify_one(
            date_of_birth=(NOW - timedelta(days=2)).date(),
            latest_outcome_at=NOW - timedelta(hours=2),
        )
        assert finding is not None and finding.code == R_RECENT

    def test_recent_outcome_beats_first_evolution(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=30),
            has_active_admission=True,
            latest_outcome_at=NOW - timedelta(hours=2),
        )
        assert finding is not None and finding.code == R_RECENT

    def test_recent_outcome_beats_residual(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_outcome_at=NOW - timedelta(hours=2),
        )
        assert finding is not None and finding.code == R_RECENT

    def test_each_rule_alone_returns_own_code(self):
        assert classify_one(
            date_of_birth=(NOW - timedelta(days=1)).date()
        ).code == R_NEWBORN
        assert classify_one(
            date_of_birth=(NOW - timedelta(days=10)).date(),
            sector=SECTOR_3A,
            sector_code=SECTOR_3A_CODE,
        ).code == R_COMPANION
        assert classify_one(
            active_admission_date=NOW - timedelta(hours=6),
            has_active_admission=True,
        ).code == R_FIRST_EVOLUTION
        assert classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
        ).code == R_RESIDUAL

    def test_at_most_one_primary_finding_per_patient(self, db):
        pront = "9000001"
        patient = make_patient(
            pront, date_of_birth=(NOW - timedelta(days=2)).date()
        )
        make_census_row(pront)
        make_recent_outcome(pront, started_at=NOW - timedelta(hours=2))
        # Patient also satisfies rule 2 (newborn): only rule 1 may win and
        # the map carries at most one finding for the prontuario.
        findings = build_bulk(
            [make_input(pront, patient_id=patient.pk)]
        )
        assert len(findings) == 1
        assert findings[pront].code == R_RECENT


# ── RED 3: exclusions (DOB, sector, events, discharge) ────────────────


class TestClassifierExclusions:
    def test_future_dob_does_not_classify_newborn(self):
        assert (
            classify_one(date_of_birth=(NOW + timedelta(days=2)).date())
            is None
        )

    def test_missing_dob_does_not_classify_newborn(self):
        assert classify_one(date_of_birth=None) is None

    def test_isolated_sector_does_not_classify_companion(self):
        assert (
            classify_one(
                date_of_birth=(NOW - timedelta(days=6)).date(),
                sector=SECTOR_OTHER,
                sector_code=SECTOR_OTHER_CODE,
            )
            is None
        )

    def test_admission_with_event_has_no_first_evolution_finding(self):
        assert (
            classify_one(
                active_admission_date=NOW - timedelta(hours=10),
                has_active_admission=True,
                active_admission_last_event_at=NOW - timedelta(hours=5),
            )
            is None
        )

    def test_admission_with_recent_event_has_no_residual(self):
        assert (
            classify_one(
                active_admission_date=NOW - timedelta(hours=72),
                has_active_admission=True,
                active_admission_last_event_at=NOW - timedelta(hours=10),
            )
            is None
        )

    def test_discharged_admission_has_no_active_rules(self):
        assert classify_one(has_active_admission=False) is None

    def test_active_admission_with_unknown_date_has_no_rules(self):
        assert (
            classify_one(
                has_active_admission=True,
                active_admission_date=None,
            )
            is None
        )

    def test_boundary_age_bands(self):
        # 0..4 days -> newborn rule; 5..28 + 3A -> companion; 29 -> none.
        assert classify_one(
            date_of_birth=(NOW - timedelta(days=4)).date()
        ).code == R_NEWBORN
        assert classify_one(
            date_of_birth=(NOW - timedelta(days=5)).date(),
            sector=SECTOR_3A,
            sector_code=SECTOR_3A_CODE,
        ).code == R_COMPANION
        assert classify_one(
            date_of_birth=(NOW - timedelta(days=28)).date(),
            sector=SECTOR_3A,
            sector_code=SECTOR_3A_CODE,
        ).code == R_COMPANION
        assert (
            classify_one(
                date_of_birth=(NOW - timedelta(days=29)).date(),
                sector=SECTOR_3A,
                sector_code=SECTOR_3A_CODE,
            )
            is None
        )


# ── RED 4: outcome superseded by later admission (auto-resolution) ────


class TestAutoResolution:
    def test_outcome_superseded_by_later_admission(self):
        assert (
            classify_one(
                latest_admission_created_at=NOW - timedelta(hours=1),
                latest_outcome_at=NOW - timedelta(hours=2),
            )
            is None
        )

    def test_outcome_holding_without_later_admission(self):
        finding = classify_one(
            latest_admission_created_at=NOW - timedelta(days=3),
            latest_outcome_at=NOW - timedelta(hours=2),
        )
        assert finding is not None and finding.code == R_RECENT

    def test_bulk_finding_disappears_when_admission_appears(self, db):
        pront = "9100001"
        patient = make_patient(pront)
        make_census_row(pront)
        make_recent_outcome(pront, started_at=NOW - timedelta(hours=2))

        inputs = [make_input(pront, patient_id=patient.pk)]
        assert build_bulk(inputs)[pront].code == R_RECENT

        # A posterior Admission resolves the recent-encounter finding on
        # the next evaluation: the obsolete code is gone (rule 4 may take
        # over as a different, current finding).
        make_active_admission(
            patient,
            admission_date=NOW - timedelta(minutes=30),
            created_at=NOW - timedelta(minutes=30),
        )
        findings = build_bulk(inputs)
        assert findings[pront].code != R_RECENT

    def test_patient_leaving_census_has_no_finding(self, db):
        in_census = "9100002"
        make_recent_outcome(
            in_census, started_at=NOW - timedelta(hours=2)
        )
        # The bulk input IS the current census: a patient absent from it
        # is never evaluated and keeps no stale entry.
        other = "9100003"
        findings = build_bulk([make_input(other)])
        assert in_census not in findings


# ── RED 5: technical axis preserved (timeout coexistence) ─────────────


@pytest.mark.django_db
class TestTechnicalAxisPreserved:
    def test_timeout_run_intact_while_finding_shown(self, admin_client):
        now = timezone.now()
        pront = "9200001"
        make_census_row(pront, captured_at=now)
        make_recent_outcome(pront, started_at=now - timedelta(hours=2))
        failed = IngestionRun.objects.create(
            status="failed",
            intent="full_sync",
            parameters_json={"patient_record": pront},
            failure_reason="timeout",
            timed_out=True,
        )

        resp = admin_client.get(reverse("services_portal:censo"))
        assert resp.status_code == 200
        assert EXPECTED[R_RECENT][0] in resp.content.decode()

        # The technical failure axis is untouched by the classifier.
        failed.refresh_from_db()
        assert failed.status == "failed"
        assert failed.timed_out is True
        assert failed.failure_reason == "timeout"

    def test_batch_status_untouched_by_classifier(self, admin_client):
        now = timezone.now()
        pront = "9200002"
        make_census_row(pront, captured_at=now)
        make_recent_outcome(pront, started_at=now - timedelta(hours=2))
        batch = CensusExecutionBatch.objects.create(status="running")

        admin_client.get(reverse("services_portal:censo"))
        batch.refresh_from_db()
        assert batch.status == "running"


# ── RED 6: residual strict conditions (DB level) ──────────────────────


@pytest.mark.django_db
class TestResidualConditions:
    def test_residual_requires_census_active_48h_no_recent_event(self):
        pront = "9300001"
        patient = make_patient(pront)
        make_census_row(pront)
        adm = make_active_admission(
            patient,
            admission_date=NOW - timedelta(hours=96),
            created_at=NOW - timedelta(days=30),
        )
        make_event(adm, happened_at=NOW - timedelta(hours=72))
        inputs = [make_input(pront, patient_id=patient.pk)]

        findings = build_bulk(inputs)
        assert findings[pront].code == R_RESIDUAL
        assert findings[pront].requires_manual_review is True
        # The label never asserts a confirmed discharge.
        assert "alta confirmada" not in findings[pront].label.lower()

        # A recent event inside the 48h window removes the finding.
        make_event(adm, happened_at=NOW - timedelta(hours=1), key="evt-2")
        assert build_bulk(inputs) == {}

    def test_admission_under_48h_is_not_residual(self):
        pront = "9300002"
        patient = make_patient(pront)
        make_census_row(pront)
        make_active_admission(
            patient,
            admission_date=NOW - timedelta(hours=10),
            created_at=NOW - timedelta(hours=10),
        )
        inputs = [make_input(pront, patient_id=patient.pk)]
        findings = build_bulk(inputs)
        assert findings[pront].code == R_FIRST_EVOLUTION

    def test_patient_not_in_census_is_never_classified(self, db):
        pront = "9300003"
        patient = make_patient(pront)
        adm = make_active_admission(
            patient,
            admission_date=NOW - timedelta(hours=96),
            created_at=NOW - timedelta(days=30),
        )
        make_event(adm, happened_at=NOW - timedelta(hours=72))
        # The input is the census: the residual patient is absent, so a
        # cohort made only of other patients carries no entry for him.
        findings = build_bulk([make_input("9300004")])
        assert pront not in findings


# ── RED 7/8: /censo HTTP badges ───────────────────────────────────────
#
# HTTP fixtures use wall-clock-relative timestamps (timezone.now())
# because the view injects timezone.now() into the classifier; the
# closed contract and priority/exclusion rules above use the
# synthetic NOW. Census rows share one captured_at per test so the
# "latest snapshot" selector sees the whole cohort.


@pytest.mark.django_db
class TestCensoFindingBadges:
    def test_recent_outcome_badge_on_desktop_and_mobile(self, admin_client):
        now = timezone.now()
        pront = "9400001"
        make_census_row(pront, captured_at=now)
        make_recent_outcome(pront, started_at=now - timedelta(hours=2))

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert resp.status_code == 200
        # Desktop <tr> and mobile card both render the same label.
        assert content.count(EXPECTED[R_RECENT][0]) >= 2

    def test_manual_review_badge_is_accessible(self, admin_client):
        now = timezone.now()
        pront = "9400002"
        make_patient(
            pront, date_of_birth=timezone.localdate() - timedelta(days=6)
        )
        make_census_row(
            pront,
            setor=SECTOR_3A,
            setor_codigo=SECTOR_3A_CODE,
            captured_at=now,
        )

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        label = EXPECTED[R_COMPANION][0]
        assert label in content
        assert content.count(label) >= 2
        # Accessible warning for requires_manual_review findings.
        assert MANUAL_REVIEW_MARK in content

    def test_newborn_badge_on_both_surfaces(self, admin_client):
        now = timezone.now()
        pront = "9400003"
        make_patient(
            pront, date_of_birth=timezone.localdate() - timedelta(days=1)
        )
        make_census_row(pront, captured_at=now)

        resp = admin_client.get(reverse("services_portal:censo"))
        assert (
            resp.content.decode().count(EXPECTED[R_NEWBORN][0]) >= 2
        )

    def test_residual_badge_on_both_surfaces(self, admin_client):
        now = timezone.now()
        pront = "9400004"
        patient = make_patient(pront)
        make_census_row(pront, captured_at=now)
        adm = make_active_admission(
            patient,
            admission_date=now - timedelta(hours=96),
            created_at=now - timedelta(days=30),
        )
        make_event(adm, happened_at=now - timedelta(hours=72))

        resp = admin_client.get(reverse("services_portal:censo"))
        assert (
            resp.content.decode().count(EXPECTED[R_RESIDUAL][0]) >= 2
        )

    def test_first_evolution_badge_on_both_surfaces(self, admin_client):
        now = timezone.now()
        pront = "9400005"
        patient = make_patient(pront)
        make_census_row(pront, captured_at=now)
        make_active_admission(
            patient,
            admission_date=now - timedelta(hours=10),
            created_at=now - timedelta(hours=10),
        )

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert content.count(EXPECTED[R_FIRST_EVOLUTION][0]) >= 2
        # Informational finding: no manual review warning.
        assert MANUAL_REVIEW_MARK not in content.split(
            EXPECTED[R_FIRST_EVOLUTION][0]
        )[1].split("</tr>")[0]

    def test_none_finding_renders_no_placeholder(self, admin_client):
        now = timezone.now()
        pront = "9400006"
        make_census_row(pront, captured_at=now)  # nothing to classify

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert resp.status_code == 200
        for label, _, _ in EXPECTED.values():
            assert label not in content
        assert "data-finding-placeholder" not in content

    def test_mixed_cohort_only_flagged_patients_have_badges(
        self, admin_client
    ):
        now = timezone.now()
        flagged = "9400007"
        plain = "9400008"
        make_census_row(flagged, captured_at=now)
        make_recent_outcome(
            flagged, started_at=now - timedelta(hours=2)
        )
        make_census_row(plain, captured_at=now)

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert EXPECTED[R_RECENT][0] in content
        # The plain patient's table row carries no finding badge.
        plain_row = next(
            line for line in content.splitlines() if plain in line
        )
        badge_line = next(
            line
            for line in content.splitlines()
            if EXPECTED[R_RECENT][0] in line
        )
        assert plain_row != badge_line


# ── RED 9: /censo contracts preserved ─────────────────────────────────


@pytest.mark.django_db
class TestCensoContractsPreserved:
    def test_anonymous_user_redirects_to_login(self, client: Client):
        resp = client.get(reverse("services_portal:censo"))
        assert resp.status_code == 302
        assert "/login/" in resp["Location"]

    def test_filters_still_work_with_findings(self, admin_client):
        now = timezone.now()
        flagged = "9500001"
        other = "9500002"
        make_census_row(flagged, captured_at=now)
        make_recent_outcome(
            flagged, started_at=now - timedelta(hours=2)
        )
        make_census_row(
            other, setor="UTI SINTEtica", setor_codigo="888",
            captured_at=now,
        )

        resp = admin_client.get(
            reverse("services_portal:censo") + "?unidade=UTI+SINTEtica"
        )
        content = resp.content.decode()
        assert resp.status_code == 200
        assert f"PACIENTE SINTEtico {other}" in content
        assert f"PACIENTE SINTEtico {flagged}" not in content

    def test_row_link_to_admission_list_preserved(self, admin_client):
        now = timezone.now()
        pront = "9500005"
        patient = make_patient(pront)
        make_census_row(pront, captured_at=now)
        make_recent_outcome(pront, started_at=now - timedelta(hours=2))

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert (
            reverse("patients:admission_list", args=[patient.pk]) in content
        )
        assert EXPECTED[R_RECENT][0] in content

    def test_export_headers_and_shape_unchanged(self, admin_client):
        now = timezone.now()
        pront = "9500006"
        patient = make_patient(pront)
        make_census_row(pront, captured_at=now)
        make_recent_outcome(pront, started_at=now - timedelta(hours=2))
        make_active_admission(
            patient,
            admission_date=now - timedelta(hours=96),
            created_at=now - timedelta(days=30),
        )

        resp = admin_client.get(reverse("services_portal:censo_export_xlsx"))
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        assert header_row == [
            "Registro", "Nome", "Setor / Unidade", "Leito",
            "Especialidade", "Data Internação", "Tempo Internação",
            "Capturado em",
        ]
        # No finding label leaks into the workbook.
        for row in ws.iter_rows(values_only=True):
            for value in row:
                for label, _, _ in EXPECTED.values():
                    assert value != label

    def test_export_skips_classifier_queries(self, admin_client):
        seed_cohort("95010", 4)

        page_url = reverse("services_portal:censo")
        export_url = reverse("services_portal:censo_export_xlsx")
        with CaptureQueriesContext(connection) as page_ctx:
            admin_client.get(page_url)
        with CaptureQueriesContext(connection) as export_ctx:
            admin_client.get(export_url)
        # The explicit option avoids classifier cost on the export path.
        assert _query_count(export_ctx) < _query_count(page_ctx)


# ── RED 10: fixed query budget (no N+1) ───────────────────────────────


@pytest.mark.django_db
class TestQueryBudget:
    def test_page_query_count_is_fixed_across_cohort_sizes(
        self, admin_client
    ):
        seed_cohort("96000", 3)
        url = reverse("services_portal:censo")
        with CaptureQueriesContext(connection) as small_ctx:
            resp = admin_client.get(url)
        assert resp.status_code == 200

        seed_cohort("97000", 9)
        with CaptureQueriesContext(connection) as big_ctx:
            resp = admin_client.get(url)
        assert resp.status_code == 200

        # Fixed tolerance of 1 query; the classifier is bulk (4 queries).
        assert abs(_query_count(big_ctx) - _query_count(small_ctx)) <= 1
        assert _query_count(big_ctx) <= 30

    def test_classifier_itself_uses_constant_queries(self):
        seed_cohort("98000", 3)
        small_inputs = [
            make_input(
                f"98000{i:02d}",
                patient_id=patient_id_of(f"98000{i:02d}"),
                sector=SECTOR_OTHER,
                sector_code=SECTOR_OTHER_CODE,
            )
            for i in range(3)
        ]
        with CaptureQueriesContext(connection) as small_ctx:
            build_bulk(small_inputs)

        seed_cohort("99000", 9)
        big_inputs = small_inputs + [
            make_input(
                f"99000{i:02d}",
                patient_id=patient_id_of(f"99000{i:02d}"),
                sector=SECTOR_OTHER,
                sector_code=SECTOR_OTHER_CODE,
            )
            for i in range(9)
        ]
        with CaptureQueriesContext(connection) as big_ctx:
            build_bulk(big_inputs)

        assert _query_count(big_ctx) == _query_count(small_ctx)
        assert _query_count(big_ctx) <= 6


# ── MSA-S1: mirror-stale admission split (rule 5 disambiguation) ──────

MIRROR_LABEL = "Suspeita de admissão órfã no espelho"
MIRROR_SEVERITY = "warning"
MIRROR_REVIEW = True


def make_movement(
    patient: Patient,
    *,
    first_seen_at: datetime,
    sector: str = SECTOR_OTHER,
) -> PatientMovement:
    """Synthesize a sector-entry ledger row (census processing output)."""
    return PatientMovement.objects.create(
        patient=patient,
        movement_date=timezone.localtime(first_seen_at).date(),
        sector=sector,
        first_seen_at=first_seen_at,
        last_seen_at=first_seen_at,
    )


def make_orphan_admission_fixture(
    pront: str, *, now: datetime
) -> Patient:
    """Rule-5 scenario: active admission >= 48h, no event in 48h."""
    patient = make_patient(pront)
    make_census_row(pront, captured_at=now)
    adm = make_active_admission(
        patient,
        admission_date=now - timedelta(hours=96),
        created_at=now - timedelta(days=30),
    )
    make_event(adm, happened_at=now - timedelta(hours=72), key=f"evt-{pront}")
    return patient


class TestMirrorStaleAdmissionClosedCode:
    def test_new_code_is_in_closed_set(self):
        assert R_MIRROR in ALL_FINDING_CODES

    def test_label_severity_review_contract(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW - timedelta(hours=1),
        )
        assert finding is not None
        assert finding.code == R_MIRROR
        assert finding.label == MIRROR_LABEL
        assert finding.severity == MIRROR_SEVERITY
        assert finding.requires_manual_review is MIRROR_REVIEW


class TestMirrorStaleAdmissionSplit:
    """Deterministic split of rule 5 by movement recency (pure fn)."""

    def test_recent_movement_yields_mirror_code(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW - timedelta(hours=1),
        )
        assert finding is not None and finding.code == R_MIRROR

    def test_movement_at_window_start_is_mirror(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW - timedelta(hours=48),
        )
        assert finding is not None and finding.code == R_MIRROR

    def test_movement_exactly_now_is_mirror(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW,
        )
        assert finding is not None and finding.code == R_MIRROR

    def test_movement_older_than_window_keeps_residual(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW - timedelta(days=3),
        )
        assert finding is not None and finding.code == R_RESIDUAL

    def test_future_movement_treated_as_absent_keeps_residual(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
            latest_movement_at=NOW + timedelta(hours=1),
        )
        assert finding is not None and finding.code == R_RESIDUAL

    def test_absent_movement_keeps_residual(self):
        finding = classify_one(
            active_admission_date=NOW - timedelta(hours=72),
            has_active_admission=True,
        )
        assert finding is not None and finding.code == R_RESIDUAL


@pytest.mark.django_db
class TestMirrorStaleAdmissionBulk:
    """Bulk service reads the movement ledger; no movement = old result."""

    def test_bulk_recent_movement_yields_mirror_stale(self):
        now = timezone.now()
        pront = "9910001"
        patient = make_orphan_admission_fixture(pront, now=now)
        make_movement(patient, first_seen_at=now - timedelta(hours=1))
        inputs = [make_input(pront, patient_id=patient.pk)]

        findings = build_patient_flow_findings(inputs, now=now)
        assert findings[pront].code == R_MIRROR
        assert findings[pront].label == MIRROR_LABEL
        assert findings[pront].severity == MIRROR_SEVERITY
        assert findings[pront].requires_manual_review is MIRROR_REVIEW

    def test_bulk_old_movement_keeps_residual(self):
        now = timezone.now()
        pront = "9910002"
        patient = make_orphan_admission_fixture(pront, now=now)
        make_movement(patient, first_seen_at=now - timedelta(days=3))
        inputs = [make_input(pront, patient_id=patient.pk)]

        findings = build_patient_flow_findings(inputs, now=now)
        assert findings[pront].code == R_RESIDUAL

    def test_bulk_future_movement_keeps_residual(self):
        now = timezone.now()
        pront = "9910003"
        patient = make_orphan_admission_fixture(pront, now=now)
        make_movement(patient, first_seen_at=now + timedelta(hours=1))
        inputs = [make_input(pront, patient_id=patient.pk)]

        findings = build_patient_flow_findings(inputs, now=now)
        assert findings[pront].code == R_RESIDUAL

    def test_bulk_without_movement_keeps_residual(self):
        now = timezone.now()
        pront = "9910004"
        patient = make_orphan_admission_fixture(pront, now=now)
        inputs = [make_input(pront, patient_id=patient.pk)]

        findings = build_patient_flow_findings(inputs, now=now)
        assert findings[pront].code == R_RESIDUAL


@pytest.mark.django_db
class TestMirrorStaleAdmissionBudget:
    def test_query_budget_with_movements_is_at_most_five(self):
        now = timezone.now()
        recent = "9910011"
        old = "9910012"
        patient_recent = make_orphan_admission_fixture(recent, now=now)
        make_movement(
            patient_recent, first_seen_at=now - timedelta(hours=1)
        )
        patient_old = make_orphan_admission_fixture(old, now=now)
        make_movement(patient_old, first_seen_at=now - timedelta(days=3))
        inputs = [
            make_input(recent, patient_id=patient_recent.pk),
            make_input(old, patient_id=patient_old.pk),
        ]

        with CaptureQueriesContext(connection) as ctx:
            findings = build_patient_flow_findings(inputs, now=now)

        # New evidence present: the split is computed from the ledger.
        assert findings[recent].code == R_MIRROR
        assert findings[old].code == R_RESIDUAL
        # Fixed budget: at most five bulk queries regardless of cohort.
        assert _query_count(ctx) <= 5


@pytest.mark.django_db
class TestMirrorStaleAdmissionCensoSurface:
    def test_censo_renders_mirror_stale_label_without_surface_changes(
        self, admin_client
    ):
        now = timezone.now()
        pront = "9910021"
        patient = make_orphan_admission_fixture(pront, now=now)
        make_movement(patient, first_seen_at=now - timedelta(hours=1))

        resp = admin_client.get(reverse("services_portal:censo"))
        content = resp.content.decode()
        assert resp.status_code == 200
        # The new closed label flows through the generic badge rendering
        # (desktop row + mobile card) with no surface modification.
        assert content.count(MIRROR_LABEL) >= 2
        # Disambiguation: the old residual label is replaced, not added.
        assert EXPECTED[R_RESIDUAL][0] not in content
