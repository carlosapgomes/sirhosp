#!/usr/bin/env python3
"""
Extrai pacientes de todos os setores do Censo via download XLSX.

Fluxo:
1) login
2) abrir Censo
3) listar setores
4) para cada setor: selecionar -> pesquisar -> exportar XLSX -> parse
5) salvar CSV consolidado

Vantagens sobre o scraping HTML antigo:
- Captura todas as 13 colunas da tabela (incluindo Dt Int e Tempo numérico)
- Sem paginação (XLS (Tudo) exporta todas as linhas)
- Mais rápido e estável
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import Frame, Page, sync_playwright

# Add parent automation/source_system so we can import shared helpers
_CURRENT_DIR = Path(__file__).resolve().parent
_SOURCE_SYSTEM_DIR = _CURRENT_DIR.parent
sys.path.insert(0, str(_SOURCE_SYSTEM_DIR))

from proxy_config import get_playwright_proxy  # noqa: E402
from source_system import aguardar_pagina_estavel, fechar_dialogos_iniciais  # noqa: E402

DEFAULT_TIMEOUT_MS = 180000

CENSO_FRAME_NAME = "i_frame_censo_diário_dos_pacientes"

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
DOWNLOADS_DIR = _PROJECT_ROOT / "downloads"
DEBUG_DIR = _PROJECT_ROOT / "debug"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrai pacientes de todos os setores via XLSX."
    )
    parser.add_argument("--headless", action="store_true", help="Executa sem UI")
    parser.add_argument("--max-setores", type=int, default=0, help="Limita setores (0 = todos)")
    parser.add_argument("--pause-ms", type=int, default=250, help="Pausa curta entre ações")
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Diretório de saída (default: <project_root>/downloads)",
    )
    parser.add_argument(
        "--csv-only",
        action="store_true",
        help="Gerar apenas CSV (não gerar JSON)",
    )
    return parser.parse_args()


def wait_visible(locator, timeout: int = 10000) -> bool:
    try:
        locator.first.wait_for(state="visible", timeout=timeout)
        return True
    except Exception:
        return False


def safe_click(locator, label: str, timeout: int = 10000) -> bool:
    if not wait_visible(locator, timeout=timeout):
        print(f"  [!] Não visível para clique: {label}")
        return False

    target = locator.first
    for force in (False, True):
        try:
            target.click(timeout=timeout, force=force)
            return True
        except Exception:
            continue

    try:
        target.evaluate("el => el.click()")
        return True
    except Exception as e:
        print(f"  [!] Clique falhou ({label}): {e}")
        return False


def get_censo_frame(page: Page, timeout_ms: int = 60000) -> Frame:
    start = time.time()
    while (time.time() - start) * 1000 < timeout_ms:
        frame = page.frame(name=CENSO_FRAME_NAME)
        if frame is not None:
            try:
                frame.locator("body").first.wait_for(state="attached", timeout=1500)
                return frame
            except Exception:
                pass
        page.wait_for_timeout(300)
    raise RuntimeError("Timeout aguardando iframe do Censo.")


def wait_ajax_idle(frame: Frame, page: Page, timeout_ms: int = 30000) -> bool:
    """Aguarda fila AJAX esvaziar e loading dialog sumir."""
    stable = 0
    deadline = time.time() + timeout_ms / 1000

    while time.time() < deadline:
        try:
            status = frame.evaluate(
                """
                () => {
                    const isVisible = (el) => {
                        if (!el) return false;
                        const style = window.getComputedStyle(el);
                        const rect = el.getBoundingClientRect();
                        const ariaHidden = (el.getAttribute('aria-hidden') || '').toLowerCase() === 'true';
                        return !ariaHidden && style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
                    };

                    const loading = document.querySelector('#form_loading');
                    let queueEmpty = true;
                    try {
                        if (window.PrimeFaces?.ajax?.Queue) {
                            if (typeof window.PrimeFaces.ajax.Queue.isEmpty === 'function') {
                                queueEmpty = window.PrimeFaces.ajax.Queue.isEmpty();
                            } else if (Array.isArray(window.PrimeFaces.ajax.Queue.requests)) {
                                queueEmpty = window.PrimeFaces.ajax.Queue.requests.length === 0;
                            }
                        }
                    } catch (_) {}

                    return {
                        queueEmpty,
                        loadingVisible: isVisible(loading),
                    };
                }
                """
            )
        except Exception:
            status = {"queueEmpty": True, "loadingVisible": False}

        if status.get("queueEmpty") and not status.get("loadingVisible"):
            stable += 1
            if stable >= 3:
                return True
        else:
            stable = 0

        page.wait_for_timeout(250)

    return False


def click_censo_icon(page: Page) -> None:
    icon = page.locator('[id="_icon_img_20323"]')
    if not safe_click(icon, "ícone do Censo", timeout=15000):
        raise RuntimeError("Não foi possível abrir o Censo.")


def extract_setores(frame: Frame, page: Page) -> list[str]:
    print("[i] Extraindo setores...")
    btn = frame.locator('[id="unidadeFuncional:unidadeFuncional:suggestion_button"]')
    if not safe_click(btn, "abrir dropdown de setores"):
        raise RuntimeError("Falha ao abrir dropdown de setores.")

    wait_ajax_idle(frame, page, timeout_ms=12000)
    page.wait_for_timeout(500)

    setores = frame.evaluate(
        """
        () => {
            const panel = document.querySelector('[id="unidadeFuncional:unidadeFuncional:suggestion_panel"]');
            const roots = panel ? [panel] : [document];
            const found = [];
            const seen = new Set();

            for (const root of roots) {
                const nodes = root.querySelectorAll('[role="cell"], td, li');
                for (const n of nodes) {
                    const t = (n.textContent || '').replace(/\\s+/g, ' ').trim();
                    if (!t || t.length < 6) continue;
                    if (!t.includes(' - ')) continue;
                    if (seen.has(t)) continue;
                    seen.add(t);
                    found.push(t);
                }
            }
            return found;
        }
        """
    )

    # Fecha dropdown clicando fora
    try:
        frame.locator("body").click(position={"x": 2, "y": 2}, timeout=1000)
    except Exception:
        pass

    return [s for s in setores if isinstance(s, str)]


def clear_setor(frame: Frame, page: Page) -> None:
    clear_btn = frame.locator('[id="unidadeFuncional:unidadeFuncional:sgClear"]')
    if clear_btn.count() > 0 and wait_visible(clear_btn, timeout=2000):
        safe_click(clear_btn, "limpar setor", timeout=4000)
        wait_ajax_idle(frame, page, timeout_ms=6000)


def select_setor(frame: Frame, page: Page, setor: str) -> bool:
    """Seleciona um setor digitando o nome no campo de autocomplete.

    Em vez de abrir o dropdown completo e caçar o item (que falha quando há
    muitos setores ou o DOM está incompleto), usa o comportamento nativo do
    PrimeFaces Autocomplete: digita o nome no campo de input, aguarda o
    filtro AJAX reduzir o painel, e clica no primeiro item resultante.
    """
    clear_setor(frame, page)

    # Localiza o input do autocomplete e garante que está visível
    input_el = frame.locator(
        '#unidadeFuncional\\:unidadeFuncional\\:suggestion_input'
    )
    if not wait_visible(input_el, timeout=8000):
        print(f"  [!] Input de autocomplete não visível para '{setor}'")
        return False

    # clear_setor pode falhar silenciosamente (botão sgClear não visível
    # ou AJAX pendente), deixando o input disabled com o valor do setor
    # anterior. Nesse caso, força a habilitação e limpeza via JS.
    if not input_el.is_enabled(timeout=3000):
        print(f"  [!] Input disabled — forçando habilitação via JS")
        frame.evaluate(
            """
            () => {
                const el = document.getElementById(
                    'unidadeFuncional:unidadeFuncional:suggestion_input'
                );
                if (el) {
                    el.removeAttribute('disabled');
                    el.classList.remove('ui-state-disabled');
                    el.value = '';
                }
            }
            """
        )
        page.wait_for_timeout(500)

    # Limpa e digita o nome completo do setor (press_sequentially
    # dispara eventos keydown/keyup que o PrimeFaces precisa para o AJAX)
    input_el.click()
    input_el.fill("")
    input_el.press_sequentially(setor, delay=50)

    # Aguarda o painel de sugestão aparecer com os resultados filtrados
    panel = frame.locator(
        '#unidadeFuncional\\:unidadeFuncional\\:suggestion_panel'
    )
    if not wait_visible(panel, timeout=12000):
        # Fallback: tenta sem o prefixo "código - " (ex: "0 T - SALA LARANJA"
        # → "SALA LARANJA"). Alguns nomes não disparam o autocomplete com o
        # formato completo.
        if ' - ' in setor:
            nome_desc = setor.split(' - ', 1)[1]
            print(
                f"  [!] Painel não apareceu com nome completo,"
                f" tentando '{nome_desc}'"
            )
            input_el.fill("")
            input_el.press_sequentially(nome_desc, delay=50)
            if not wait_visible(panel, timeout=10000):
                print(f"  [!] Painel de sugestão não apareceu para '{setor}'")
                return False
        else:
            print(f"  [!] Painel de sugestão não apareceu para '{setor}'")
            return False

    page.wait_for_timeout(500)

    # Clica no primeiro (e idealmente único) item do painel
    item = panel.locator('a, li, div[role="cell"], td').first
    if item.count() == 0:
        print(f"  [!] Nenhum item no painel para '{setor}'")
        return False

    try:
        item.click(timeout=5000)
    except Exception:
        # Último recurso: JS click
        try:
            item.evaluate("el => el.click()")
        except Exception as e:
            print(f"  [!] Clique no item falhou para '{setor}': {e}")
            return False

    wait_ajax_idle(frame, page, timeout_ms=15000)
    return True


def click_pesquisar(frame: Frame) -> bool:
    btn = frame.locator('[id="bt_pesquisar:button"]')
    return safe_click(btn, "botão Pesquisar", timeout=12000)


def get_current_setor_info(frame: Frame) -> dict[str, str]:
    """Extract current sector code and full name from the results page."""
    return frame.evaluate(
        """
        () => {
            const input = document.querySelector('#unidadeFuncional\\\\:unidadeFuncional\\\\:suggestion_input');
            const label = document.querySelector('#unidadeFuncional\\\\:unidadeFuncional\\\\:sgDescricaoLabel');
            return {
                codigo: (input?.value || '').trim(),
                nome: (label?.textContent || '').replace(/\\s+/g, ' ').trim(),
            };
        }
        """
    )


# ── XLSX export helpers ─────────────────────────────────────────────


def _click_export_button(frame: Frame, page: Page) -> None:
    """Click the 'Exportar para Arquivo' button using JS.

    Playwright's click() hangs on this PrimeFaces commandButton because
    it triggers an AJAX event that Playwright misinterprets.  Using
    element.click() via JS dispatch avoids the deadlock.

    In sectors with many patients the button is below the fold,
    so we scroll to the bottom of the page first.
    """
    # Aguarda o botão Exportar estar presente no DOM (o paginator pode
    # renderizar após o tbody em datatables muito longos como este).
    try:
        frame.locator('[title="Exportar para Arquivo"]').wait_for(
            state='attached', timeout=15000
        )
    except Exception:
        raise RuntimeError("Botão Exportar não apareceu no DOM após 15s")
    page.wait_for_timeout(300)

    # Scroll agressivo até o final da tabela (setores com 60+ pacientes
    # deixam o botão muito abaixo do fold visível).  window.scrollTo no
    # body do iframe não é suficiente porque o scroll real está no
    # wrapper .ui-datatable-tablewrapper do PrimeFaces.
    frame.evaluate(
        """
        () => {
            // Rola wrappers de datatable até o fim
            const wrappers = document.querySelectorAll(
                '.ui-datatable-tablewrapper, .ui-datatable-scrollable-body'
            );
            for (const w of wrappers) {
                w.scrollTop = w.scrollHeight;
            }
            // Rola também o body e o html
            window.scrollTo(0, document.body.scrollHeight);
            document.documentElement.scrollTop = document.documentElement.scrollHeight;
        }
        """
    )
    page.wait_for_timeout(200)

    # Tecla End como reforço (alcança o fundo mesmo em tabelas muito longas)
    frame.locator('body').press('End')
    page.wait_for_timeout(200)

    # PageDown repetido como último recurso (para datatables que
    # renderizam linhas sob demanda conforme scroll)
    for _ in range(3):
        frame.locator('body').press('PageDown')
        page.wait_for_timeout(150)

    clicked = frame.evaluate(
        """
        () => {
            const btn = document.querySelector(
                '[title="Exportar para Arquivo"]'
            );
            if (!btn) return false;
            btn.scrollIntoView({ block: 'center' });
            btn.click();
            return true;
        }
        """
    )
    if not clicked:
        raise RuntimeError("Não foi possível clicar em Exportar para Arquivo.")
    page.wait_for_timeout(300)


def _click_xls_tudo(frame: Frame, page: Page) -> None:
    """Click the 'XLS (Tudo)' menu item using JS."""
    clicked = frame.evaluate(
        """
        () => {
            const all = document.querySelectorAll('a');
            for (const a of all) {
                if (a.offsetParent === null) continue;
                const t = (a.textContent || '').replace(/\\s+/g, ' ').trim().toLowerCase();
                if (t === 'xls (tudo)') {
                    a.scrollIntoView({ block: 'center' });
                    a.click();
                    return true;
                }
            }
            return false;
        }
        """
    )
    if not clicked:
        raise RuntimeError("Não foi possível clicar em XLS (Tudo).")


def export_setor_xlsx(frame: Frame, page: Page) -> Path:
    """Clica em Exportar para Arquivo → XLS (Tudo) e retorna o path do XLSX baixado.

    Usa JS clicks em vez de Playwright clicks para evitar deadlock com
    eventos AJAX do PrimeFaces. Copia o arquivo para um path estável
    antes de retornar (evita race condition com tempfile do Playwright).
    """
    import shutil

    _click_export_button(frame, page)
    page.wait_for_timeout(500)

    # Prepara o download e clica em "XLS (Tudo)"
    with page.expect_download(timeout=120000) as download_info:
        _click_xls_tudo(frame, page)

    download = download_info.value
    download_path = download.path()
    if download_path is None:
        raise RuntimeError("Download não produziu arquivo no disco.")

    # Copy to a stable path (Playwright may clean up the temp file)
    stable_path = download_path.with_suffix(".stable.xlsx")
    shutil.copy2(download_path, stable_path)

    print(f"    [i] XLSX baixado: {download.suggested_filename} ({stable_path.stat().st_size} bytes)")
    return stable_path


def parse_setor_xlsx(
    xlsx_path: Path,
    setor_codigo: str,
    setor_nome: str,
) -> list[dict[str, Any]]:
    """Parseia um XLSX exportado do AGHU e retorna lista de pacientes.

    Mapeamento de colunas do AGHU:
        Qrt/Leito → qrt_leito
        Prontuário → prontuario
        Nome/Situação → nome
        Idade → idade
        Dt Int → dt_int (DD/MM)
        Esp → esp
        Médico → medico
        Dt Mvt → dt_mvt
        Alta → alta
        Origem → origem
        Tempo → tempo (numérico)
        Convênio → convenio
        Transferência → transferencia
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError("XLSX sem sheet ativa.")

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows_iter)]

    # Build column index mapping
    col_map: dict[str, int] = {}
    col_aliases = {
        "Qrt/Leito": "qrt_leito",
        "Prontuário": "prontuario",
        "Prontuario": "prontuario",
        "Nome/Situação": "nome",
        "Nome": "nome",
        "Idade": "idade",
        "Dt Int": "dt_int",
        "Esp": "esp",
        "Médico": "medico",
        "Dt Mvt": "dt_mvt",
        "Alta": "alta",
        "Origem": "origem",
        "Tempo": "tempo",
        "Convênio": "convenio",
        "ConvÃªnio": "convenio",
        "Transferência": "transferencia",
    }

    for i, h in enumerate(headers):
        if h in col_aliases:
            col_map[col_aliases[h]] = i

    required = {"qrt_leito", "prontuario", "nome"}
    missing = required - set(col_map.keys())
    if missing:
        print(f"    [A] Colunas encontradas: {headers}")
        raise ValueError(f"Colunas obrigatórias ausentes no XLSX: {missing}")

    def _sanitize_prontuario(raw: Any) -> str:
        """Sanitize prontuário: convert number to int→str, remove separators.

        openpyxl returns numeric cells as int or float. A cell with
        19017680.0 becomes "19017680.0" via str(), which breaks the
        downstream worker that receives --patient-record '19017680.0'.
        """
        if raw is None:
            return ""
        if isinstance(raw, (int, float)):
            raw = str(int(raw))
        else:
            raw = str(raw).strip()
        # Remove any remaining dots, slashes, or non-digit separators
        return raw.replace(".", "").replace("/", "").strip()

    pacientes: list[dict[str, Any]] = []
    for row in rows_iter:
        prontuario = _sanitize_prontuario(
            row[col_map["prontuario"]] if col_map["prontuario"] < len(row) else None
        )
        nome = str(row[col_map["nome"]]).strip() if col_map["nome"] < len(row) and row[col_map["nome"]] is not None else ""

        if not prontuario and not nome:
            continue

        qrt_leito = str(row[col_map["qrt_leito"]]).strip() if col_map["qrt_leito"] < len(row) and row[col_map["qrt_leito"]] is not None else ""

        dt_int = ""
        if "dt_int" in col_map and col_map["dt_int"] < len(row) and row[col_map["dt_int"]] is not None:
            dt_int = str(row[col_map["dt_int"]]).strip()

        tempo = None
        if "tempo" in col_map and col_map["tempo"] < len(row) and row[col_map["tempo"]] is not None:
            val = row[col_map["tempo"]]
            try:
                tempo = int(float(val))
            except (ValueError, TypeError):
                pass

        esp = str(row[col_map["esp"]]).strip() if col_map["esp"] < len(row) and row[col_map["esp"]] is not None else ""

        idade = ""
        if "idade" in col_map and col_map["idade"] < len(row) and row[col_map["idade"]] is not None:
            idade = str(int(float(row[col_map["idade"]]))) if isinstance(row[col_map["idade"]], (int, float)) else str(row[col_map["idade"]]).strip()

        convenio = ""
        if "convenio" in col_map and col_map["convenio"] < len(row) and row[col_map["convenio"]] is not None:
            convenio = str(row[col_map["convenio"]]).strip()

        dt_mvt = ""
        if "dt_mvt" in col_map and col_map["dt_mvt"] < len(row) and row[col_map["dt_mvt"]] is not None:
            dt_mvt = str(row[col_map["dt_mvt"]]).strip()

        alta = ""
        if "alta" in col_map and col_map["alta"] < len(row) and row[col_map["alta"]] is not None:
            alta = str(row[col_map["alta"]]).strip()

        origem = ""
        if "origem" in col_map and col_map["origem"] < len(row) and row[col_map["origem"]] is not None:
            origem = str(row[col_map["origem"]]).strip()

        pacientes.append({
            "setor_codigo": setor_codigo,
            "setor": setor_nome,
            "qrt_leito": qrt_leito,
            "prontuario": prontuario,
            "nome": nome,
            "esp": esp,
            "dt_int": dt_int,
            "tempo": tempo,
            "idade": idade,
            "convenio": convenio,
            "dt_mvt": dt_mvt,
            "alta": alta,
            "origem": origem,
        })

    wb.close()
    return pacientes


# ── Output helpers ──────────────────────────────────────────────────


def save_results(results: list[dict], csv_only: bool = False) -> tuple[Path | None, Path]:
    DOWNLOADS_DIR.mkdir(exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")

    json_path = None
    csv_path = DOWNLOADS_DIR / f"censo-todos-pacientes-xlsx-{ts}.csv"

    if not csv_only:
        json_path = DOWNLOADS_DIR / f"censo-todos-pacientes-xlsx-{ts}.json"
        json_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "setor_codigo", "setor", "qrt_leito", "prontuario", "nome", "esp",
            "dt_int", "tempo", "idade", "convenio", "dt_mvt", "alta", "origem",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for entry in results:
            setor = entry.get("setor", "")
            for p in entry.get("pacientes", []):
                row = {
                    "setor_codigo": entry.get("setor_codigo", ""),
                    "setor": setor,
                    "qrt_leito": p.get("qrt_leito", ""),
                    "prontuario": p.get("prontuario", ""),
                    "nome": p.get("nome", ""),
                    "esp": p.get("esp", ""),
                    "dt_int": p.get("dt_int", ""),
                    "tempo": p.get("tempo", ""),
                    "idade": p.get("idade", ""),
                    "convenio": p.get("convenio", ""),
                    "dt_mvt": p.get("dt_mvt", ""),
                    "alta": p.get("alta", ""),
                    "origem": p.get("origem", ""),
                }
                w.writerow(row)

    return json_path, csv_path


def save_debug(page: Page) -> None:
    DEBUG_DIR.mkdir(exist_ok=True)
    ts = time.strftime("%Y%m%d-%H%M%S")
    try:
        page.screenshot(path=str(DEBUG_DIR / f"censo-xlsx-{ts}.png"), full_page=True)
        (DEBUG_DIR / f"censo-xlsx-{ts}.html").write_text(page.content(), encoding="utf-8")
        frame = page.frame(name=CENSO_FRAME_NAME)
        if frame:
            (DEBUG_DIR / f"censo-xlsx-{ts}-iframe.html").write_text(frame.content(), encoding="utf-8")
        print(f"[i] Debug salvo em debug/censo-xlsx-{ts}.*")
    except Exception as e:
        print(f"[!] Falha ao salvar debug: {e}")


# ── Main flow ───────────────────────────────────────────────────────


def run(
    source_system_url: str,
    username: str,
    password: str,
    headless: bool,
    max_setores: int,
    pause_ms: int,
    csv_only: bool = False,
) -> None:
    with sync_playwright() as pw:
        browser = context = page = None
        try:
            print("[i] Abrindo navegador...")
            _proxy = get_playwright_proxy()
            _launch_kwargs: dict[str, Any] = {
                "headless": headless,
                "args": ["--ignore-certificate-errors"],
            }
            if _proxy:
                _launch_kwargs["proxy"] = _proxy
            browser = pw.chromium.launch(**_launch_kwargs)
            context = browser.new_context(ignore_https_errors=True)
            page = context.new_page()
            page.set_default_timeout(DEFAULT_TIMEOUT_MS)

            print(f"[i] Acessando {source_system_url}...")
            page.goto(source_system_url)

            print("[i] Login...")
            page.get_by_role("textbox", name="Nome de usuário").fill(username)
            page.get_by_role("textbox", name="Senha").fill(password)
            page.get_by_role("button", name="Entrar").click()
            aguardar_pagina_estavel(page)

            print("[i] Fechando diálogos iniciais...")
            fechar_dialogos_iniciais(page)

            print("[i] Abrindo Censo...")
            click_censo_icon(page)
            frame = get_censo_frame(page)
            wait_ajax_idle(frame, page, timeout_ms=15000)

            setores = extract_setores(frame, page)
            if not setores:
                raise RuntimeError("Nenhum setor encontrado.")

            if max_setores > 0:
                setores = setores[:max_setores]

            print(f"[i] Total de setores a processar: {len(setores)}")

            results: list[dict] = []
            _retried: list[str] = []
            _failed: list[str] = []

            for i, setor in enumerate(setores, start=1):
                print(f"\n[{i}/{len(setores)}] {setor}")

                _max_retries = 2
                _last_error = None

                for _attempt in range(_max_retries + 1):
                    if _attempt > 0:
                        print(
                            f"    [RETRY {_attempt}/{_max_retries}]"
                            f" após erro: {_last_error}"
                        )
                        page.wait_for_timeout(3000)

                    try:
                        if not select_setor(frame, page, setor):
                            raise RuntimeError(
                                "não conseguiu selecionar setor"
                            )

                        page.wait_for_timeout(pause_ms)

                        setor_codigo = ""
                        setor_nome_completo = setor

                        if not click_pesquisar(frame):
                            raise RuntimeError("falha no clique de pesquisar")

                        # Extrair código e nome completo após pesquisar
                        setor_info = get_current_setor_info(frame)
                        setor_codigo = setor_info.get("codigo", "")
                        setor_nome_completo = (
                            setor_info.get("nome") or setor
                        )

                        # Aguarda a tabela carregar completamente
                        wait_ajax_idle(frame, page, timeout_ms=60000)

                        # Aguarda o tbody ter conteúdo real
                        try:
                            frame.wait_for_function(
                                """
                                () => {
                                    const tbody = document.querySelector(
                                        '[id$="resultList_data"]'
                                    );
                                    if (!tbody) return false;
                                    return (
                                        (
                                            tbody.textContent || ''
                                        ).trim().length > 0
                                    );
                                }
                                """,
                                timeout=45000,
                            )
                        except Exception:
                            page.wait_for_timeout(3000)

                        # Verifica se o setor tem pacientes
                        tem_pacientes = frame.evaluate(
                            """
                            () => {
                                const tbody = document.querySelector(
                                    '[id$="resultList_data"]'
                                );
                                if (!tbody) return false;
                                const txt = (
                                    tbody.textContent || ''
                                ).trim();
                                if (txt.includes('Nenhum registro'))
                                    return false;
                                const rows = tbody.querySelectorAll('tr');
                                if (rows.length === 0) return false;
                                for (const row of rows) {
                                    const td = row.querySelector('td');
                                    if (
                                        td
                                        && td.textContent.trim().length > 0
                                    ) {
                                        return true;
                                    }
                                }
                                return false;
                            }
                            """
                        )
                        if not tem_pacientes:
                            print(f"    setor vazio — pulando")
                            results.append({
                                "setor_codigo": setor_codigo,
                                "setor": setor_nome_completo,
                                "pacientes": [],
                            })
                            break  # sai do loop de retry

                        # Exporta XLSX
                        xlsx_path = export_setor_xlsx(frame, page)

                        # Parseia o XLSX
                        pacientes = parse_setor_xlsx(
                            xlsx_path, setor_codigo, setor_nome_completo
                        )
                        print(
                            f"    pacientes extraídos: {len(pacientes)}"
                        )

                        results.append({
                            "setor_codigo": setor_codigo,
                            "setor": setor_nome_completo,
                            "pacientes": pacientes,
                        })
                        if _attempt > 0:
                            _retried.append(setor)
                        break  # sucesso — sai do loop de retry

                    except Exception as _exc:
                        _last_error = _exc
                        if _attempt < _max_retries:
                            continue
                        _failed.append(setor)
                        print(f"    [ERRO] {_last_error}")
                        results.append({
                            "setor_codigo": "",
                            "setor": setor,
                            "pacientes": [],
                            "erro": str(_last_error),
                        })

            json_path, csv_path = save_results(results, csv_only=csv_only)

            total_pacientes = sum(len(r.get("pacientes", [])) for r in results)
            setores_erro = sum(1 for r in results if r.get("erro"))
            print("\n" + "=" * 70)
            print(f"Setores processados: {len(results)}")
            print(f"Setores com erro:   {setores_erro}")
            if _retried:
                print(
                    f"Setores recuperados após retry:"
                    f" {len(_retried)}"
                )
                for s in _retried:
                    print(f"  ✓ {s}")
            if _failed:
                print(
                    f"Setores com falha definitiva:"
                    f" {len(_failed)}"
                )
                for s in _failed:
                    print(f"  ✗ {s}")
            print(f"Total pacientes:    {total_pacientes}")
            if json_path:
                print(f"JSON: {json_path}")
            print(f"CSV:  {csv_path}")
            print("=" * 70)

        except Exception as e:
            print(f"[FATAL] {e}")
            if page is not None:
                save_debug(page)
            raise
        finally:
            if context:
                context.close()
            if browser:
                browser.close()
            pw.stop()


def main() -> None:
    args = parse_args()

    if args.output_dir:
        global DOWNLOADS_DIR  # noqa: PLW0603
        DOWNLOADS_DIR = Path(args.output_dir)

    run(
        source_system_url=os.getenv("SOURCE_SYSTEM_URL", ""),
        username=os.getenv("SOURCE_SYSTEM_USERNAME", ""),
        password=os.getenv("SOURCE_SYSTEM_PASSWORD", ""),
        headless=args.headless,
        max_setores=args.max_setores,
        pause_ms=args.pause_ms,
        csv_only=args.csv_only,
    )


if __name__ == "__main__":
    main()
